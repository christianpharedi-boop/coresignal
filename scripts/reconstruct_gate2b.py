"""Reconstruct Gate 2B event and event-pair populations from the verified workbook only."""
from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_tables(path: Path) -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    event_sheet = workbook["Extended Data Table 1"]
    pair_sheet = workbook["Extended Data Table 2"]
    events = []
    for row in event_sheet.iter_rows(min_row=3, values_only=True):
        date_value, index_value = row[:2]
        if date_value is None and index_value is None:
            continue
        events.append({"event_id": int(index_value), "origin": str(date_value)})
    pairs = []
    for row in pair_sheet.iter_rows(min_row=4, values_only=True):
        pair_id, quality, ym, im, event_a, event_b, date_a, date_b, latitude, longitude, multiplet, reference = row[:12]
        if not isinstance(pair_id, str) or not pair_id.strip().startswith("P") or event_a is None or event_b is None:
            continue
        pairs.append({
            "pair_id": str(pair_id).strip(),
            "event_a": int(event_a),
            "event_b": int(event_b),
            "date_a": str(date_a),
            "date_b": str(date_b),
            "latitude": latitude,
            "longitude": longitude,
            "multiplet": None if multiplet is None else str(multiplet),
            "reference": None if reference is None else str(reference),
        })
    return events, pairs


def reconstruct(workbook_path: Path) -> dict:
    events, pairs = load_tables(workbook_path)
    event_ids = [event["event_id"] for event in events]
    pair_ids = [pair["pair_id"] for pair in pairs]
    event_id_set = set(event_ids)
    dangling = [pair["pair_id"] for pair in pairs if pair["event_a"] not in event_id_set or pair["event_b"] not in event_id_set]
    duplicate_events = sorted({value for value in event_ids if event_ids.count(value) > 1})
    duplicate_pairs = sorted({value for value in pair_ids if pair_ids.count(value) > 1})
    dates = sorted(event["origin"] for event in events)
    return {
        "gate": "M1-Gate-2B",
        "status": "DEFER_RECONSTRUCTION",
        "lod_accessed": False,
        "waveform_accessed": False,
        "input": {
            "path": str(workbook_path),
            "filename": workbook_path.name,
            "sha256": sha256_file(workbook_path),
        },
        "event_reconstruction": {
            "expected_count": 121,
            "observed_count": len(events),
            "unique_id_count": len(set(event_ids)),
            "first_origin": dates[0] if dates else None,
            "last_origin": dates[-1] if dates else None,
            "duplicate_event_ids": duplicate_events,
            "pass": len(events) == 121 and len(set(event_ids)) == 121,
        },
        "pair_reconstruction": {
            "paper_narrative_count": 143,
            "table_label_count": 142,
            "narrative_expected_count": 143,
            "observed_count": len(pairs),
            "unique_pair_id_count": len(set(pair_ids)),
            "duplicate_pair_ids": duplicate_pairs,
            "dangling_event_references": dangling,
            "referential_integrity_pass": not dangling,
            "count_discrepancy": "workbook title states 142; workbook contains 143 pair records P01-P143",
            "count_discrepancy_requires_resolution": len(pairs) != 143,
            "pass": len(pairs) == 143 and len(set(pair_ids)) == 143 and not dangling,
        },
        "station_reconstruction": {
            "required_arrays": ["ILAR", "YKA"],
            "status": "PENDING_QUERY_EXECUTION",
            "pass": False,
        },
        "waveform_retrieval": {
            "authorized": False,
            "reason": "station reconstruction and archive feasibility remain unresolved",
        },
        "decision": "DEFER_RECONSTRUCTION",
        "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    report = reconstruct(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
