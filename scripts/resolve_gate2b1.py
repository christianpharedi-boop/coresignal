"""Resolve Gate 2B.1 station metadata and date-specific feasibility without waveforms or LOD."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


STATION_QUERY_URL = "https://service.iris.edu/fdsnws/station/1/query?station={station}&level=channel&format=text"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_station(path: Path) -> dict:
    with path.open(encoding="utf-8") as handle:
        rows = list(csv.reader(handle, delimiter="|"))
    header = [item.strip().lstrip("#") for item in rows[0]]
    values = [item.strip() for item in rows[1]]
    data = dict(zip(header, values))
    return {
        "network": data["Network"],
        "station": data["Station"],
        "location": data["Location"],
        "channel": data["Channel"],
        "latitude": float(data["Latitude"]),
        "longitude": float(data["Longitude"]),
        "elevation": float(data["Elevation"]),
        "start_date": data["StartTime"],
        "end_date": data["EndTime"] or None,
        "sensor_description": data["SensorDescription"],
        "scale": data["Scale"],
        "scale_frequency": data["ScaleFreq"],
        "scale_units": data["ScaleUnits"],
        "sample_rate": float(data["SampleRate"]),
        "response_metadata_identity": path.name,
        "response_sha256": sha256_file(path),
        "archive_provider": "EarthScope IRIS / NSF SAGE",
        "query_url": STATION_QUERY_URL.format(station=data["Station"]),
        "retrieval_timestamp_utc": "2026-08-17T18:11:11Z",
    }


def load_pairs(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Extended Data Table 2"]
    pairs = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        pair_id, _, _, _, event_a, event_b, date_a, date_b = row[:8]
        if not isinstance(pair_id, str) or not pair_id.strip().startswith("P") or event_a is None or event_b is None:
            continue
        pairs.append({
            "pair_id": pair_id.strip(),
            "event_a": int(event_a),
            "event_b": int(event_b),
            "event_a_origin_utc": str(date_a),
            "event_b_origin_utc": str(date_b),
        })
    return pairs


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--station-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    inventory = [
        parse_station(args.station_dir / "IM_ILAR_FB_SHZ_station.txt"),
        parse_station(args.station_dir / "IM_YKA_FB_SHZ_station.txt"),
    ]
    pairs = load_pairs(args.workbook)
    feasibility = []
    for pair in pairs:
        for station in inventory:
            station_start = datetime.fromisoformat(station["start_date"].replace("Z", "+00:00"))
            event_a = datetime.fromisoformat(pair["event_a_origin_utc"].replace("Z", "+00:00"))
            event_b = datetime.fromisoformat(pair["event_b_origin_utc"].replace("Z", "+00:00"))
            operational_a = event_a >= station_start
            operational_b = event_b >= station_start
            feasibility.append({
                **pair,
                "station": station["station"],
                "network": station["network"],
                "location": station["location"],
                "channel": station["channel"],
                "station_operational_at_event_a": operational_a,
                "station_operational_at_event_b": operational_b,
                "channel_available_at_event_a": operational_a,
                "channel_available_at_event_b": operational_b,
                "archive_coverage_status": "UNKNOWN",
                "pkikp_request_feasible": "DEFERRED",
                "decision_basis": "station metadata covers event dates; archive-level waveform availability was not established",
            })

    report = {
        "gate": "M1-Gate-2B.1",
        "status": "DEFER_RECONSTRUCTION",
        "lod_accessed": False,
        "waveform_bytes_accessed": False,
        "station_inventory": inventory,
        "pair_count": len(pairs),
        "station_pair_feasibility_count": len(feasibility),
        "pair_feasibility": feasibility,
        "archive_coverage_query": {
            "status": "UNAVAILABLE",
            "endpoint": "https://service.iris.edu/fdsnws/availability/1/query",
            "observed_response": "HTTP 410 Gone",
            "interpretation": "No archive coverage claim made from this failed endpoint.",
        },
        "waveform_retrieval_authorized": False,
        "waveform_retrieval_authorization_basis": "archive coverage remains unknown; request set is not frozen",
        "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
